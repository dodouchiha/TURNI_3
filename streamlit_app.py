import streamlit as st
import pandas as pd
import calendar
from datetime import datetime
import holidays
from io import BytesIO
import requests
import json
import base64
import time
import logging
import tempfile
import os
from functools import wraps
from openpyxl.styles import PatternFill, Font, Alignment
import re # Importato per valida_nome_medico

# --- CONFIGURAZIONE LOGGING ---
def setup_logging():
    """Configura logging per debug e monitoraggio"""
    # Rimuovi handler esistenti se la funzione viene chiamata più volte (es. in Streamlit)
    # Questo previene la duplicazione dei log se lo script viene rieseguito
    root_logger = logging.getLogger()
    for handler in root_logger.handlers[:]:
        root_logger.removeHandler(handler)
        handler.close()

    log_file_path = 'medical_shifts.log'
    # Controlla se il file di log è scrivibile, altrimenti logga solo su stream
    log_handlers = [logging.StreamHandler()]
    try:
        # Tenta di aprire in modalità append per verificare i permessi
        with open(log_file_path, 'a'): 
            pass
        log_handlers.append(logging.FileHandler(log_file_path, mode='a', encoding='utf-8'))
    except PermissionError:
        print(f"Attenzione: Permesso negato per scrivere su {log_file_path}. Il logging su file è disabilitato.")
    except Exception as e:
        print(f"Attenzione: Errore durante l'inizializzazione del logging su file ({log_file_path}): {e}. Il logging su file è disabilitato.")

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=log_handlers
    )
    return logging.getLogger(__name__)

logger = setup_logging()
logger.info("Applicazione Gestione Turni Medici avviata.")

# --- CONFIGURAZIONE INIZIALE ---
st.set_page_config(page_title="Gestione Turni Medici", layout="wide", initial_sidebar_state="expanded")

# --- COSTANTI ---
COL_DATA = "Data"
COL_GIORNO = "Giorno"
COL_FESTIVO = "Festivo" 
COL_NOME_FESTIVO = "Nome Festivo"
TIPI_ASSENZA = ["Presente", "Ferie", "Malattia", "Congresso", "Lezione", "Altro"]
MEDICI_BACKUP_FILE = "medici_backup.json"

# --- CONFIGURAZIONE GITHUB ---
GITHUB_USER = st.secrets.get("GITHUB_USER", "dodouchiha") # Default se non in secrets
REPO_NAME = st.secrets.get("REPO_NAME", "turni_3")     # Default se non in secrets
FILE_PATH = st.secrets.get("FILE_PATH_MEDICI", "medici.json") # Default se non in secrets
GITHUB_TOKEN = st.secrets.get("GITHUB_TOKEN")

API_URL = f"https://api.github.com/repos/{GITHUB_USER}/{REPO_NAME}/contents/{FILE_PATH}"
HEADERS = {"Authorization": f"token {GITHUB_TOKEN}", "Accept": "application/vnd.github.v3+json"}

# --- FUNZIONI DI VALIDAZIONE ---
def valida_nome_medico(nome):
    """Valida il nome del medico inserito"""
    nome = nome.strip()
    if not nome or len(nome) < 2:
        return False, "Il nome deve contenere almeno 2 caratteri."
    if len(nome) > 50:
        return False, "Il nome non può superare 50 caratteri."
    if not re.match(r"^[a-zA-ZÀ-ÿ\s.'-]+$", nome): # Aggiunto il trattino
        return False, "Il nome può contenere solo lettere, spazi, punti, apostrofi e trattini."
    return True, ""

def verifica_configurazione():
    """Verifica la configurazione prima dell'avvio."""
    issues = []
    logger.info("Verifica configurazione...")
    
    if not GITHUB_TOKEN:
        issues.append("Token GitHub (GITHUB_TOKEN) mancante nelle secrets.")
    if not GITHUB_USER:
        issues.append("Utente GitHub (GITHUB_USER) mancante nelle secrets o non impostato.")
    if not REPO_NAME:
        issues.append("Nome Repository (REPO_NAME) mancante nelle secrets o non impostato.")
    if not FILE_PATH:
        issues.append("Percorso File Medici (FILE_PATH_MEDICI) mancante nelle secrets o non impostato.")

    if GITHUB_TOKEN and GITHUB_USER and REPO_NAME: # Prova la connessione solo se i dati base ci sono
        try:
            test_response = requests.get(
                f"https://api.github.com/repos/{GITHUB_USER}/{REPO_NAME}", 
                headers=HEADERS, 
                timeout=10
            )
            if test_response.status_code == 404:
                issues.append(f"Repository GitHub '{GITHUB_USER}/{REPO_NAME}' non trovato o non accessibile.")
            elif test_response.status_code == 401:
                issues.append("Token GitHub non valido o con permessi insufficienti per il repository.")
            elif test_response.status_code != 200:
                issues.append(f"Errore connessione GitHub al repository: {test_response.status_code}.")
            else:
                logger.info("Connessione al repository GitHub riuscita.")
        except requests.exceptions.RequestException as e:
            issues.append(f"Impossibile connettersi a GitHub: {str(e)}.")
    else:
        issues.append("Dati di base per la connessione a GitHub mancanti.")
    
    if issues:
        logger.warning(f"Problemi di configurazione rilevati: {issues}")
    else:
        logger.info("Configurazione verificata con successo.")
    return issues

# --- BACKUP LOCALE ---
def salva_backup_locale(data, filename):
    """Salva backup locale in caso di problemi GitHub."""
    try:
        # Usa una directory specifica dell'app se possibile, altrimenti tempdir
        # In ambienti come Streamlit Cloud, tempfile.gettempdir() è più affidabile
        # per la persistenza limitata tra le sessioni (ma non garantita a lungo termine).
        # Per un vero backup persistente locale, bisognerebbe montare un volume.
        backup_dir = os.path.join(tempfile.gettempdir(), "medical_shifts_app_backup")
        os.makedirs(backup_dir, exist_ok=True)
        
        backup_path = os.path.join(backup_dir, filename)
        with open(backup_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Backup locale salvato: {backup_path}")
        return backup_path
    except Exception as e:
        logger.warning(f"Impossibile salvare backup locale '{filename}': {e}")
        st.sidebar.warning(f"⚠️ Impossibile salvare backup locale.") # Feedback più discreto
        return None

def carica_backup_locale(filename):
    """Carica backup locale se disponibile."""
    try:
        backup_dir = os.path.join(tempfile.gettempdir(), "medical_shifts_app_backup")
        backup_path = os.path.join(backup_dir, filename)
        
        if os.path.exists(backup_path):
            with open(backup_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            logger.info(f"Backup locale '{filename}' caricato con successo da: {backup_path}")
            return data
        else:
            logger.info(f"Nessun backup locale trovato per '{filename}' in {backup_path}")
    except Exception as e:
        logger.warning(f"Errore caricamento backup locale '{filename}': {e}")
    return None

# --- RETRY DECORATOR ---
def retry_github_api(max_retries=3, delay_seconds=2):
    """Decorator per retry automatico delle chiamate GitHub API con backoff esponenziale."""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            last_exception = None
            for attempt in range(max_retries):
                try:
                    return func(*args, **kwargs)
                except requests.exceptions.HTTPError as e:
                    last_exception = e
                    # Non fare retry per 404 (Not Found) o 401 (Unauthorized) perché di solito non sono temporanei
                    if e.response.status_code in [404, 401]:
                        logger.error(f"Errore HTTP GitHub non recuperabile ({e.response.status_code}) per {func.__name__}. Nessun ulteriore tentativo.")
                        raise 
                    
                    if e.response.status_code == 403 and "rate limit" in e.response.text.lower():
                        # Specifica gestione del rate limit
                        if attempt < max_retries - 1:
                            # GitHub spesso fornisce l'header 'Retry-After'
                            retry_after = int(e.response.headers.get("Retry-After", delay_seconds * (2 ** attempt)))
                            logger.warning(f"Rate limit GitHub per {func.__name__}. Tentativo {attempt + 1}/{max_retries}. Riprovo tra {retry_after}s...")
                            st.sidebar.warning(f"⏳ Rate limit GitHub. Riprovo tra {retry_after}s...")
                            time.sleep(retry_after)
                            continue
                    elif e.response.status_code >= 500:  # Errori Server GitHub
                        if attempt < max_retries - 1:
                            wait_time = delay_seconds * (2 ** attempt)
                            logger.warning(f"Errore server GitHub ({e.response.status_code}) per {func.__name__}. Tentativo {attempt + 1}/{max_retries}. Riprovo tra {wait_time}s...")
                            st.sidebar.warning(f"⏳ Errore server GitHub. Riprovo tra {wait_time}s...")
                            time.sleep(wait_time)
                            continue
                    logger.error(f"Errore HTTP GitHub ({e.response.status_code}) per {func.__name__}: {e.response.text}")
                    raise # Solleva l'eccezione se non è gestita specificamente sopra o se i tentativi sono esauriti
                except requests.exceptions.RequestException as e: # Errori di connessione, timeout, etc.
                    last_exception = e
                    if attempt < max_retries - 1:
                        wait_time = delay_seconds * (2 ** attempt)
                        logger.warning(f"Errore di connessione per {func.__name__}. Tentativo {attempt + 1}/{max_retries}. Riprovo tra {wait_time}s... Errore: {e}")
                        st.sidebar.warning(f"⏳ Errore connessione. Riprovo tra {wait_time}s...")
                        time.sleep(wait_time)
                        continue
                    logger.error(f"Errore di connessione finale per {func.__name__} dopo {max_retries} tentativi: {e}")
                    raise
            # Se tutti i tentativi falliscono, solleva l'ultima eccezione catturata
            if last_exception:
                raise last_exception
            return None # Teoricamente non raggiungibile se un'eccezione viene sempre sollevata
        return wrapper
    return decorator

# --- FUNZIONI GITHUB MIGLIORATE ---
@retry_github_api()
def carica_medici_da_github():
    logger.info(f"Tentativo caricamento medici da GitHub: {API_URL}")
    res = requests.get(API_URL, headers=HEADERS, timeout=15)
    res.raise_for_status() # Solleverà HTTPError per status codes 4xx/5xx
    
    contenuto = res.json()
    if "content" not in contenuto or "sha" not in contenuto:
        logger.error("Risposta GitHub per carica_medici non contiene 'content' o 'sha'.")
        raise ValueError("Formato risposta GitHub inatteso.")
        
    file_sha = contenuto["sha"]
    elenco_json = base64.b64decode(contenuto["content"]).decode('utf-8')
    elenco = json.loads(elenco_json)
    
    st.session_state.sha_medici = file_sha
    logger.info(f"Medici caricati da GitHub: {len(elenco)} elementi. SHA: {file_sha}")
    return elenco

def inizializza_elenco_medici():
    """Tenta di caricare i medici da GitHub, poi da backup locale, altrimenti lista vuota."""
    try:
        elenco = carica_medici_da_github()
        salva_backup_locale(elenco, MEDICI_BACKUP_FILE) # Salva backup dopo caricamento GitHub riuscito
        return elenco
    except requests.exceptions.HTTPError as e_gh:
        if e_gh.response.status_code == 404:
            st.sidebar.warning(f"File '{FILE_PATH}' non trovato su GitHub. Sarà creato al primo salvataggio.")
            logger.info("File medici non trovato su GitHub (404), inizializzazione vuota.")
            st.session_state.sha_medici = None # Importante per la creazione del file
            return []
        else: # Altri errori HTTP da GitHub
            logger.error(f"Errore HTTP caricamento medici da GitHub: {e_gh.response.status_code}. Tento backup locale.")
            st.sidebar.error("⚠️ Errore GitHub nel caricare i medici.")
            backup_data = carica_backup_locale(MEDICI_BACKUP_FILE)
            if backup_data is not None:
                st.sidebar.warning("Medici caricati da backup locale.")
                # Non abbiamo uno SHA valido da GitHub in questo caso, il prossimo salvataggio potrebbe fallire
                # se il file esiste su GitHub e lo SHA è cambiato. 
                # Considerare di impostare sha_medici = None per forzare una potenziale sovrascrittura (rischioso)
                # o informare l'utente. Per ora, lo lasciamo così.
                return backup_data
            else: # Backup fallito o non esistente
                st.sidebar.error("Impossibile caricare medici da GitHub o da backup locale.")
                return [] # Ritorna lista vuota per permettere all'app di avviarsi in stato degradato
    except requests.exceptions.RequestException as e_req: # Errori di rete
        logger.error(f"Errore di rete caricamento medici da GitHub: {e_req}. Tento backup locale.")
        st.sidebar.error("⚠️ Errore di rete nel caricare i medici.")
        backup_data = carica_backup_locale(MEDICI_BACKUP_FILE)
        if backup_data is not None:
            st.sidebar.warning("Medici caricati da backup locale.")
            return backup_data
        else:
            st.sidebar.error("Impossibile caricare medici da GitHub o da backup locale.")
            return []
    except Exception as e_gen: # Altri errori imprevisti
        logger.error(f"Errore imprevisto caricamento elenco medici: {e_gen}. Tento backup locale.")
        st.sidebar.error(f"⚠️ Errore imprevisto: {e_gen}")
        backup_data = carica_backup_locale(MEDICI_BACKUP_FILE)
        if backup_data is not None:
            st.sidebar.warning("Medici caricati da backup locale.")
            return backup_data
        else:
            st.sidebar.error("Impossibile caricare medici da GitHub o da backup locale.")
            return []

@retry_github_api()
def salva_medici_su_github(lista_medici, sha_corrente):
    if not isinstance(lista_medici, list):
        logger.error("Tentativo di salvare 'lista_medici' che non è una lista.")
        raise ValueError("lista_medici deve essere una lista.")
        
    blob = json.dumps(lista_medici, indent=2, ensure_ascii=False).encode('utf-8')
    encoded_content = base64.b64encode(blob).decode('utf-8')
    
    data = {
        "message": f"Aggiornamento elenco medici - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "content": encoded_content,
        "branch": "main" # Assicurati che il branch sia corretto
    }
    
    if sha_corrente: # Se è un aggiornamento di un file esistente
        data["sha"] = sha_corrente
    # Se sha_corrente è None, si sta tentando di creare un nuovo file.
    # GitHub gestisce questo implicitamente: se non c'è SHA e il file non esiste, lo crea.
    # Se non c'è SHA e il file esiste, darà errore (a meno che l'API non sia cambiata).
    # Per essere più espliciti, se sha_corrente è None e si sa che il file potrebbe esistere,
    # bisognerebbe prima fare una GET per ottenere lo SHA o confermare il 404.
    # La logica attuale di carica_medici imposta sha_medici = None se il file è 404.

    logger.info(f"Tentativo salvataggio {len(lista_medici)} medici su GitHub. SHA usato: {sha_corrente}")
    res = requests.put(API_URL, headers=HEADERS, json=data, timeout=15)
    res.raise_for_status() # Solleverà HTTPError per status codes 4xx/5xx
    
    if res.status_code in [200, 201]: # 200 OK (update), 201 Created (new file)
        nuovo_sha = res.json()["content"]["sha"]
        st.session_state.sha_medici = nuovo_sha
        logger.info(f"Medici salvati su GitHub con successo. Nuovo SHA: {nuovo_sha}")
        salva_backup_locale(lista_medici, MEDICI_BACKUP_FILE) # Aggiorna backup dopo salvataggio GitHub
        return True
    
    logger.warning(f"Salvataggio medici su GitHub: status code inaspettato {res.status_code}. Risposta: {res.text}")
    return False

# --- VERIFICA CONFIGURAZIONE ALL'AVVIO ---
if 'config_checked' not in st.session_state:
    config_issues = verifica_configurazione()
    if config_issues:
        st.error("⚠️ **Problemi di Configurazione dell'Applicazione Rilevati:**")
        for issue in config_issues:
            st.error(f"   • {issue}")
        st.warning("L'applicazione potrebbe non funzionare come previsto. Controlla le secrets e le impostazioni del repository.")
        if any("Token GitHub mancante" in issue for issue in config_issues) or \
           any("Dati di base per la connessione a GitHub mancanti" in issue for issue in config_issues) :
            logger.critical("Configurazione GitHub essenziale mancante. Arresto dell'applicazione.")
            st.info("Per supporto, contatta l'amministratore.")
            st.stop()
    st.session_state.config_checked = True

# --- INIZIALIZZAZIONE STATO ---
if 'elenco_medici_completo' not in st.session_state:
    with st.spinner("Caricamento elenco medici..."):
        st.session_state.elenco_medici_completo = inizializza_elenco_medici()
        
if 'sha_medici' not in st.session_state: # Lo SHA potrebbe essere stato impostato da inizializza_elenco_medici
    st.session_state.sha_medici = None

# --- FUNZIONI CALENDARIO MIGLIORATE ---
@st.cache_data(ttl=3600, show_spinner="Generazione calendario...") # Cache per 1 ora
def genera_calendario_cached(anno, mese, medici_serialized_sorted):
    """Genera calendario con cache per migliorare performance.
       medici_serialized_sorted deve essere una stringa JSON serializzata di una lista ordinata.
    """
    try:
        medici_list = json.loads(medici_serialized_sorted)
        logger.info(f"Cache miss o scaduta: Generazione calendario per {mese}/{anno} con {len(medici_list)} medici")
        return genera_struttura_calendario(anno, mese, medici_list)
    except Exception as e:
        logger.error(f"Errore generazione calendario (cached): {e}")
        st.error(f"Impossibile generare il calendario: {e}")
        return pd.DataFrame()

def genera_struttura_calendario(anno, mese, medici_selezionati):
    """Genera la struttura base del calendario mensile."""
    try:
        _, ultimo_giorno = calendar.monthrange(anno, mese)
        date_del_mese = pd.date_range(
            start=f"{anno}-{mese:02d}-01", 
            end=f"{anno}-{mese:02d}-{ultimo_giorno}"
        )
        
        try:
            festivita_anno = holidays.country_holidays("IT", years=anno)
        except Exception as e_hol: # Più generico per coprire altri errori della libreria holidays
            logger.warning(f"Festività per l'anno {anno} non disponibili o errore libreria holidays: {e_hol}")
            st.sidebar.caption(f"⚠️ Festività per {anno} non caricate.") # Feedback discreto
            festivita_anno = {}
        
        df_cols = {
            COL_DATA: date_del_mese,
            COL_GIORNO: [d.strftime("%A") for d in date_del_mese],
            COL_FESTIVO: [d.date() in festivita_anno for d in date_del_mese],
            COL_NOME_FESTIVO: [festivita_anno.get(d.date(), "") for d in date_del_mese]
        }
        
        for medico in medici_selezionati:
            df_cols[medico] = "Presente"
        
        df = pd.DataFrame(df_cols)
        logger.info(f"Calendario per {mese}/{anno} generato: {len(df)} giorni, {len(medici_selezionati)} medici.")
        return df
        
    except Exception as e:
        logger.error(f"Errore grave durante la generazione della struttura del calendario: {e}")
        st.error(f"Errore critico nella generazione del calendario: {e}")
        return pd.DataFrame()

def aggiorna_calendario_se_necessario(anno, mese, medici_pianificati_lista):
    """Centralizza la logica di aggiornamento calendario usando st.session_state."""
    try:
        # Per la chiave di cache e confronto, serializza sempre la lista ordinata dei medici
        medici_serialized_sorted = json.dumps(sorted(list(set(medici_pianificati_lista))), ensure_ascii=False)
        
        # Una chiave più semplice per il session_state, hash non è necessario qui
        # se usiamo una rappresentazione canonica (stringa json ordinata).
        current_config_representation = f"{anno}-{mese}-{medici_serialized_sorted}"
        
        if (st.session_state.get('last_calendar_config_representation') != current_config_representation or 
            'df_turni' not in st.session_state or 
            st.session_state.df_turni is None): # Aggiunto controllo per df_turni nullo
            
            if medici_pianificati_lista: # Solo se ci sono medici da pianificare
                # Passa la lista (non serializzata) alla funzione che usa la cache, 
                # la cache stessa gestirà la serializzazione interna se necessario o la usa così.
                # Per @st.cache_data, gli argomenti devono essere hashable. Le liste non lo sono.
                # Quindi passiamo la stringa serializzata ordinata.
                st.session_state.df_turni = genera_calendario_cached(
                    anno, mese, medici_serialized_sorted
                )
            else:
                st.session_state.df_turni = pd.DataFrame() # Nessun medico, calendario vuoto
            
            st.session_state.last_calendar_config_representation = current_config_representation
            if st.session_state.df_turni is not None and not st.session_state.df_turni.empty:
                 logger.info(f"Calendario aggiornato/ricaricato per: Anno {anno}, Mese {mese}, Medici: {len(medici_pianificati_lista)}")
            elif not medici_pianificati_lista:
                 logger.info(f"Nessun medico selezionato. Calendario impostato a vuoto.")
            else: # df_turni è None o vuoto nonostante ci fossero medici
                 logger.warning(f"Calendario per {anno}-{mese} risultante vuoto o None nonostante {len(medici_pianificati_lista)} medici.")
                 
    except Exception as e:
        logger.error(f"Errore critico durante l'aggiornamento del calendario: {e}")
        st.error(f"Impossibile aggiornare il calendario: {e}")
        st.session_state.df_turni = pd.DataFrame() # Fallback a DataFrame vuoto

# --- EXPORT MIGLIORATO ---
def esporta_con_formattazione(df_originale, nome_file_base):
    """Export Excel con formattazione migliorata."""
    if df_originale is None or df_originale.empty:
        logger.warning("Tentativo di esportare un DataFrame vuoto o nullo.")
        st.error("Nessun dato valido da esportare.")
        return None
    try:
        output = BytesIO()
        df_export = df_originale.copy()

        # Conversione colonna Data se necessario e formattazione per Excel
        if COL_DATA in df_export.columns:
            # Assicura che sia datetime
            if not pd.api.types.is_datetime64_any_dtype(df_export[COL_DATA]):
                try: 
                    df_export[COL_DATA] = pd.to_datetime(df_export[COL_DATA])
                except Exception as e_conv:
                    logger.warning(f"Impossibile convertire colonna '{COL_DATA}' in datetime per Excel: {e_conv}")
            # Applica formattazione specifica per date in Excel (se la colonna è datetime)
            if pd.api.types.is_datetime64_any_dtype(df_export[COL_DATA]):
                 df_export[COL_DATA] = df_export[COL_DATA].dt.date # Usa solo la parte data per evitare problemi di timezone in Excel

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='Turni')
            worksheet = writer.sheets['Turni']
            
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Blu scuro
            header_font = Font(color="FFFFFF", bold=True, name='Calibri', size=11)
            cell_font = Font(name='Calibri', size=10)
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            for col_idx, column_name in enumerate(df_export.columns, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_col=len(df_export.columns), max_row=worksheet.max_row), 2):
                for cell in row:
                    cell.font = cell_font
                    if isinstance(cell.value, (datetime, pd.Timestamp, datetime.date)): # Se è una dataa
                        cell.number_format = 'DD/MM/YYYY' # Formato data Excel
                        cell.alignment = center_alignment
                    elif isinstance(cell.value, str) and cell.column_letter == worksheet.cell(row=row_idx, column=df_export.columns.get_loc(COL_GIORNO) + 1).column_letter: # Giorno
                        cell.alignment = left_alignment
                    elif isinstance(cell.value, str) and cell.column_letter == worksheet.cell(row=row_idx, column=df_export.columns.get_loc(COL_NOME_FESTIVO) + 1).column_letter: # Nome Festivo
                        cell.alignment = left_alignment
                    else: # Altre colonne (tipi assenza, nomi medici)
                        cell.alignment = center_alignment
            
            # Auto-adjust column widths
            for column_cells in worksheet.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if cell.value:
                            # Per le date, considera la lunghezza del formato stringa
                            if isinstance(cell.value, (datetime, pd.Timestamp, datetime.date)):
                                value_to_measure = cell.value.strftime('%d/%m/%Y')
                            else:
                                value_to_measure = str(cell.value)
                            max_length = max(max_length, len(value_to_measure))
                    except: pass
                adjusted_width = min(max(max_length + 2, 12), 40) # Min 12, Max 40
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        logger.info(f"Export Excel '{nome_file_base}' generato con successo.")
        return output
        
    except Exception as e:
        logger.error(f"Errore durante la generazione dell'export Excel '{nome_file_base}': {e}", exc_info=True)
        st.error(f"Errore nella generazione del file Excel: {e}")
        return None

# --- UI SIDEBAR ---
st.sidebar.title("🗓️ Gestione Turni")
st.sidebar.markdown("App per la pianificazione dei turni medici.")
st.sidebar.divider()

st.sidebar.header("👨‍⚕️ Medici")
with st.sidebar.form("form_aggiungi_medico", clear_on_submit=True):
    nuovo_medico_input = st.text_input("➕ Nome nuovo medico (es. Rossi Mario)").strip()
    submitted_add = st.form_submit_button("Aggiungi Medico", type="primary")

if submitted_add and nuovo_medico_input:
    valido, messaggio = valida_nome_medico(nuovo_medico_input)
    if not valido:
        st.sidebar.error(messaggio)
    elif nuovo_medico_input not in st.session_state.elenco_medici_completo:
        elenco_aggiornato = st.session_state.elenco_medici_completo + [nuovo_medico_input]
        elenco_aggiornato.sort()
        
        with st.spinner("Salvataggio medico su GitHub..."):
            try:
                if salva_medici_su_github(elenco_aggiornato, st.session_state.get("sha_medici")):
                    st.session_state.elenco_medici_completo = elenco_aggiornato # Aggiorna stato locale solo se GitHub OK
                    st.toast(f"Medico '{nuovo_medico_input}' aggiunto!", icon="✅")
                    logger.info(f"Medico aggiunto e salvato su GitHub: {nuovo_medico_input}")
                    st.rerun()
                else: # Salvataggio fallito, l'errore è già loggato e mostrato da salva_medici_su_github
                    st.sidebar.error("❌ Salvataggio su GitHub non riuscito. Controlla i log o riprova.")
            except Exception as e_save: # Cattura eccezioni non gestite da salva_medici_su_github
                 logger.error(f"Eccezione non gestita durante il salvataggio del medico: {e_save}", exc_info=True)
                 st.sidebar.error(f"❌ Errore critico salvataggio: {e_save}")
    else:
        st.sidebar.warning(f"'{nuovo_medico_input}' è già presente nell'elenco.")

if st.session_state.elenco_medici_completo:
    options_rimuovi = ["--- Seleziona per rimuovere ---"] + sorted(list(set(st.session_state.elenco_medici_completo)))
    current_selection_rimuovi = st.session_state.get("medico_da_rimuovere_selection", options_rimuovi[0])
    
    try:
        default_index_rimuovi = options_rimuovi.index(current_selection_rimuovi)
    except ValueError: default_index_rimuovi = 0
    
    medico_da_rimuovere = st.sidebar.selectbox(
        "🗑️ Rimuovi medico", options=options_rimuovi, index=default_index_rimuovi,
        key="selectbox_rimuovi_medico_key"
    )
    st.session_state.medico_da_rimuovere_selection = medico_da_rimuovere
    
    if medico_da_rimuovere != options_rimuovi[0] and st.sidebar.button("Conferma Rimozione", key="button_conferma_rimozione", type="secondary"):
        medici_temp = st.session_state.elenco_medici_completo.copy()
        medici_temp.remove(medico_da_rimuovere)
        
        with st.spinner(f"Rimozione di '{medico_da_rimuovere}' da GitHub..."):
            try:
                if salva_medici_su_github(medici_temp, st.session_state.get("sha_medici")):
                    st.session_state.elenco_medici_completo = medici_temp
                    st.toast(f"Medico '{medico_da_rimuovere}' rimosso.", icon="🗑️")
                    logger.info(f"Medico rimosso e salvato su GitHub: {medico_da_rimuovere}")
                    
                    if 'medici_pianificati' in st.session_state and medico_da_rimuovere in st.session_state.medici_pianificati:
                        st.session_state.medici_pianificati.remove(medico_da_rimuovere)
                    if 'df_turni' in st.session_state: del st.session_state.df_turni
                    st.session_state.medico_da_rimuovere_selection = options_rimuovi[0]
                    st.rerun()
                else:
                    st.sidebar.error("❌ Rimozione da GitHub non riuscita.")
            except Exception as e_remove:
                 logger.error(f"Eccezione non gestita durante la rimozione del medico: {e_remove}", exc_info=True)
                 st.sidebar.error(f"❌ Errore critico rimozione: {e_remove}")
else:
    st.sidebar.caption("Nessun medico presente nell'elenco.")

st.sidebar.divider()

# Selezione medici per pianificazione
st.sidebar.header("🎯 Pianificazione")
default_medici_pianificati = st.session_state.get('medici_pianificati', [])
# Assicura che i medici di default siano ancora validi e presenti nell'elenco completo
valid_default_medici = [m for m in default_medici_pianificati if m in st.session_state.elenco_medici_completo]
# Se non ci sono default validi ma ci sono medici, seleziona tutti i medici
if not valid_default_medici and st.session_state.elenco_medici_completo:
    valid_default_medici = st.session_state.elenco_medici_completo[:] # Usa una copia

medici_pianificati = st.sidebar.multiselect(
    "👨‍⚕️ Medici da includere nel piano:",
    options=sorted(list(set(st.session_state.elenco_medici_completo))), # Unici e ordinati
    default=valid_default_medici, 
    key="multiselect_medici_pianificati",
    help="Seleziona i medici per cui generare il calendario dei turni."
)

if 'medici_pianificati' not in st.session_state or set(st.session_state.medici_pianificati) != set(medici_pianificati):
    st.session_state.medici_pianificati = medici_pianificati
    # Forza l'aggiornamento del calendario se la lista dei medici pianificati cambia
    aggiorna_calendario_se_necessario(st.session_state.get('selected_anno', datetime.today().year), 
                                     st.session_state.get('selected_mese', datetime.today().month), 
                                     medici_pianificati)


# --- SELEZIONE MESE E ANNO ---
st.sidebar.header("🗓️ Periodo")
oggi = datetime.today()
# Mantiene la selezione precedente o imposta default
idx_mese_default = st.session_state.get('selected_mese_index', oggi.month - 1)
# Assicura che l'indice sia valido (es. se la sessione viene resettata malamente)
if not 0 <= idx_mese_default < 12: idx_mese_default = oggi.month - 1

anni_disponibili = list(range(oggi.year - 3, oggi.year + 4)) # Range di 7 anni
idx_anno_default = st.session_state.get('selected_anno_index', anni_disponibili.index(oggi.year))
if not 0 <= idx_anno_default < len(anni_disponibili): idx_anno_default = anni_disponibili.index(oggi.year)


col1_sidebar, col2_sidebar = st.sidebar.columns(2)
lista_mesi = list(range(1, 13))
selected_mese = col1_sidebar.selectbox(
    "Mese:", lista_mesi, index=idx_mese_default,
    format_func=lambda x: calendar.month_name[x], key="selectbox_mese"
)
selected_anno = col2_sidebar.selectbox(
    "Anno:", anni_disponibili, index=idx_anno_default, key="selectbox_anno"
)

# Aggiorna lo stato della sessione per persistere le selezioni di mese/anno
# e forza l'aggiornamento del calendario se cambiano
if st.session_state.get('selected_mese_index') != lista_mesi.index(selected_mese) or \
   st.session_state.get('selected_anno_index') != anni_disponibili.index(selected_anno):
    st.session_state.selected_mese_index = lista_mesi.index(selected_mese)
    st.session_state.selected_anno_index = anni_disponibili.index(selected_anno)
    st.session_state.selected_mese = selected_mese # Salva anche il valore, non solo l'indice
    st.session_state.selected_anno = selected_anno
    aggiorna_calendario_se_necessario(selected_anno, selected_mese, medici_pianificati)
else: # Se mese/anno non sono cambiati, ma il df_turni non esiste, rigeneralo
    if 'df_turni' not in st.session_state or st.session_state.df_turni is None:
         aggiorna_calendario_se_necessario(selected_anno, selected_mese, medici_pianificati)


nome_mese_corrente = calendar.month_name[selected_mese]

# --- FUNZIONE DI STYLING ---
def evidenzia_weekend_festivi(row_series):
    """Evidenzia weekend e festivi nel DataFrame."""
    try:
        data_val = row_series[COL_DATA] # pd.Timestamp
        # COL_FESTIVO è booleano, direttamente dalla riga
        is_weekend = data_val.weekday() >= 5 # 5: Sabato, 6: Domenica
        
        # Stile di default per cella
        style_array = [''] * len(row_series)
        
        if is_weekend or row_series[COL_FESTIVO]:
            # Applica a tutte le celle della riga
            style_array = ['background-color: #e9ecef'] * len(row_series) # Grigio chiaro Bootstrap-like
        
        # Esempio: colorare diversamente le assenze
        # for i, col_name in enumerate(row_series.index):
        #     if col_name in st.session_state.get('medici_pianificati', []) and row_series[col_name] != "Presente":
        #         style_array[i] = f"{style_array[i]}; color: #007bff; font-style: italic;" # Blu per assenze

        return style_array
    except Exception as e_style:
        logger.warning(f"Errore durante lo styling della riga: {e_style}. Riga: {row_series.to_dict() if isinstance(row_series, pd.Series) else row_series}")
        return [''] * len(row_series) # Fallback a nessuno stile

# --- AREA PRINCIPALE DELL'APP ---
st.title(f"🗓️ Pianificazione Turni Medici")
st.markdown(f"### {nome_mese_corrente} {selected_anno}")

if not medici_pianificati:
    st.info("👈 **Nessun medico selezionato per la pianificazione.** Scegli almeno un medico dalla sidebar per iniziare.")
elif 'df_turni' not in st.session_state or st.session_state.df_turni is None or st.session_state.df_turni.empty:
    # Questo stato potrebbe indicare un errore nella generazione del calendario o nessun medico selezionato
    st.warning("📅 Il calendario è vuoto. Verifica le selezioni o prova a ricaricare la pagina.")
    logger.warning(f"df_turni è vuoto o None. Medici pianificati: {len(medici_pianificati)}. Stato sessione df_turni: {st.session_state.get('df_turni')}")
else:
    # df_turni esiste e non è vuoto
    st.markdown("#### ✨ **Visualizzazione Calendario**")
    df_visualizzazione = st.session_state.df_turni.copy()
    
    try:
        styled_df = df_visualizzazione.style \
            .apply(evidenzia_weekend_festivi, axis=1) \
            .format({COL_DATA: lambda dt: dt.strftime('%d/%m/%Y (%a)') if pd.notna(dt) else ""})
        
        # Nascondi la colonna booleana COL_FESTIVO dalla visualizzazione
        styled_df = styled_df.hide([COL_FESTIVO], axis='columns') 
        
        st.dataframe(styled_df, use_container_width=True, hide_index=True, height=(len(df_visualizzazione) + 1) * 35 + 3) # Altezza dinamica
    except Exception as e_df_display:
        logger.error(f"Errore durante la visualizzazione del DataFrame stilizzato: {e_df_display}", exc_info=True)
        st.error("⚠️ Errore nella visualizzazione del calendario. Mostro dati grezzi.")
        st.dataframe(df_visualizzazione, use_container_width=True, hide_index=True) # Fallback
    
    st.divider()

    st.markdown("#### 📝 **Inserisci/Modifica Assenze**")
    # Colonne per l'editor: Data e Giorno per contesto, più le colonne dei medici.
    cols_per_editor = [COL_DATA, COL_GIORNO] + medici_pianificati
    
    column_config_editor = {
        COL_DATA: st.column_config.DateColumn("Data", format="DD/MM/YYYY", disabled=True, width="small"),
        COL_GIORNO: st.column_config.TextColumn("Giorno", disabled=True, width="small"),
    }
    
    for medico in medici_pianificati:
        nome_cognome = medico.split()
        nome_display = nome_cognome[-1] if len(nome_cognome) > 1 else medico # Tenta di usare il cognome
        nome_display = nome_display.capitalize() # Capitalizza
        column_config_editor[medico] = st.column_config.SelectboxColumn(
            f"Dr. {nome_display}",
            help=f"Stato di servizio per {medico}", 
            options=TIPI_ASSENZA, 
            required=True, # Assicura che ogni cella abbia un valore
            width="medium" 
        )
    
    # Assicurati che df_turni abbia le colonne necessarie prima di creare df_editor_input
    df_editor_input_valid = True
    for col_edit in cols_per_editor:
        if col_edit not in st.session_state.df_turni.columns:
            logger.error(f"Colonna '{col_edit}' mancante in df_turni per l'editor. Colonne disponibili: {st.session_state.df_turni.columns.tolist()}")
            st.error(f"Errore interno: colonna '{col_edit}' non trovata per l'editor. Riprova o contatta supporto.")
            df_editor_input_valid = False
            break
            
    if df_editor_input_valid:
        df_editor_input = st.session_state.df_turni[cols_per_editor].copy()
        
        editor_key = f"data_editor_assenze_{st.session_state.get('last_calendar_config_representation', 'default_key')}"
        
        try:
            edited_df_assenze = st.data_editor(
                df_editor_input,
                column_config=column_config_editor,
                use_container_width=True,
                hide_index=True,
                num_rows="fixed", # Il numero di giorni nel mese è fisso
                key=editor_key,
                height=(len(df_editor_input) + 1) * 35 + 3 # Altezza dinamica per l'editor
            )

            modifiche_effettuate_editor = False
            for medico_col in medici_pianificati: # Itera solo sulle colonne dei medici
                # Confronta solo se entrambe le serie esistono e sono dello stesso tipo
                if (medico_col in st.session_state.df_turni.columns and 
                    medico_col in edited_df_assenze.columns and
                    st.session_state.df_turni[medico_col].dtype == edited_df_assenze[medico_col].dtype):
                    
                    if not st.session_state.df_turni[medico_col].equals(edited_df_assenze[medico_col]):
                        st.session_state.df_turni[medico_col] = edited_df_assenze[medico_col].copy() # Salva una copia
                        modifiche_effettuate_editor = True
            
            if modifiche_effettuate_editor:
                st.toast("Assenze nel calendario aggiornate localmente.", icon="📝")
                logger.info("Assenze modificate dall'utente e aggiornate in st.session_state.df_turni.")
                # Non serve un rerun qui, perché la tabella stilizzata sopra leggerà da st.session_state.df_turni
                # alla prossima interazione o riesecuzione automatica di Streamlit.
                # Per forzare l'aggiornamento immediato della tabella stilizzata, si potrebbe fare un rerun,
                # ma potrebbe essere un'esperienza utente interrotta.
                # st.rerun() # Valutare se necessario per UX
                
        except Exception as e_data_editor:
            logger.error(f"Errore con st.data_editor: {e_data_editor}", exc_info=True)
            st.error("⚠️ Si è verificato un errore nell'editor delle assenze. Prova a ricaricare la pagina.")

    st.divider()

    # --- ESPORTAZIONE ---
    st.markdown("#### 📤 **Esporta Calendario**")
    
    if 'df_turni' in st.session_state and st.session_state.df_turni is not None and not st.session_state.df_turni.empty:
        col_export_btn, col_export_info = st.columns([0.3, 0.7]) # Più spazio per le info
        
        nome_file_excel = f"Turni_{nome_mese_corrente.replace(' ', '_')}_{selected_anno}.xlsx"

        with col_export_btn:
            # Genera i dati per il download solo quando il bottone è premuto per efficienza
            # Tuttavia, st.download_button richiede 'data' al momento della creazione del bottone.
            # Quindi, generiamo i dati prima.
            excel_export_data = esporta_con_formattazione(
                st.session_state.df_turni.copy(), # Passa una copia per evitare modifiche accidentali
                nome_file_excel
            )
            
            if excel_export_data:
                st.download_button(
                    label="📥 Scarica Excel",
                    data=excel_export_data,
                    file_name=nome_file_excel,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel_button",
                    help="Scarica il calendario corrente in formato Excel con stili applicati."
                )
            else:
                st.error("Errore nella preparazione del file Excel per il download.")
        
        with col_export_info:
            st.caption(f"""
                Clicca per scaricare il file '{nome_file_excel}'. 
                Il file includerà tutti i medici e le assenze attualmente visualizzate, 
                con formattazione ottimizzata.
            """)
    else:
        st.caption("Nessun dato disponibile da esportare al momento.")

st.sidebar.divider()
st.sidebar.markdown(f"""
<div style="font-size: 0.8em; text-align: center; color: grey;">
    Gestione Turni Medici v0.9<br>
    {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
</div>
""", unsafe_allow_html=True)

logger.info("Rendering della pagina completato.")
